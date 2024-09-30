import itertools
import logging
import re
from collections import defaultdict
from functools import lru_cache, partial
from typing import TYPE_CHECKING, Callable, Hashable, List, Optional, Tuple

from pycountry import countries
from sdmx.model import common, v21

if TYPE_CHECKING:
    import pathlib

    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

log = logging.getLogger(__name__)

#: Concepts and metadata attributes in the TDC metadata structure.
CONCEPTS = {
    "DATAFLOW": (
        "Data flow ID",
        """A unique identifier for the data flow (=data source, data set, etc.).

We suggest to use IDs like ‘VN001’, where ‘VN’ is the ISO 3166 alpha-2 country
code, and ‘001’ is a unique number. The value MUST match the name of the sheet
in which it appears.""",
    ),
    "DATA_PROVIDER": (
        "Data provider",
        """Organization or individual that provides the data and any related metadata.

This can be as general (“IEA”) or specific (organization unit/department, specific
person responsible, contact details, etc.) as appropriate.""",
    ),
    "URL": (
        "URL or web address",
        "Location on the Internet with further information about the data flow.",
    ),
    "MEASURE": (
        "Measure (‘indicator’)",
        """Statistical concept for which data are provided in the data flow.

If the data flow contains data for multiple measures, give each one separated by
semicolons. Example: “Number of cars; passengers per vehicle”.

This SHOULD NOT duplicate the value for ‘UNIT_MEASURE’. Example: “Annual driving
distance per vehicle”, not “Kilometres per vehicle”.""",
    ),
    "UNIT_MEASURE": (
        "Unit of measure",
        """Unit in which the data values are expressed.

If ‘MEASURE’ contains 2+ items separated by semicolons, give the respective units in the
same way and order. If there are no units, write ‘dimensionless’, ‘1’, or similar.""",
    ),
    "DIMENSION": (
        "Dimensions",
        """Formally, the “statistical concept used in combination with other statistical
concepts to identify a statistical series or individual observations.”

Record all dimensions of the data, either in a bulleted or numbered list, or
separated by semicolons. In parentheses, give some indication of the scope
and/or resolution of the data along each dimension. Most data have at least time
and space dimensions.

Example:

- TIME_PERIOD (annual, 5 years up to 2021)
- REF_AREA (whole country; VN only)
- Vehicle type (12 different types: […])
- Emissions species (CO2 and 4 others)""",
    ),
    "DATA_DESCR": (
        "Data description",
        """Any information about the data flow that does not fit in other attributes.

Until or unless other metadata attributes are added to this metadata structure/
template, this MAY include:

- Any conditions on data access, e.g. publicly available, proprietary, fee or
  subscription required, available on request, etc.
- Frequency of data updates.
- Any indication of quality, including third-party references that indicate data
  quality.
""",
    ),
    "COMMENT": (
        "Comment",
        """Any other information about the metadata values, for instance discrepancies or
unclear or missing information.

Precede comments with initials; append to existing comments to keep
chronological order; and include a date (for example, “2024-07-24”) if helpful.""",
    ),
}

#: README text for the TDC metadata file format.
README_TEXT = """This file is an unofficial, prototype TDC format for metadata.
loosely imitates the Eurostat format. These files contain metadata (information
*about* data) based on the SDMX information model, but their layout (sheet
names, columns, etc.) is not specified by the SDMX standard, hence ‘unofficial’.

This file has the following sheets.

README
======

This sheet.

Attributes
==========

- One row per metadata attribute (or 'field').
- Columns for the name; description; and ID (short and machine-readable) of each
  attribute. See these descriptions to learn what to write for each attribute.

One or more additional sheets
=============================

- The name (or title) of each sheet corresponds to the identity (ID) of the data
  flow that is described by the metadata in that sheet.
- In Column A, the name of the metadata attribute. Each name MUST exactly
  match one appearing in the "Attributes" sheet. Some names MAY be omitted.
- In Column B, the actual metadata. These may be empty.

TEMPLATE
========

To add information about additional data flows not included in existing sheets
(above), you can copy and rename this sheet.
"""


def _header(ws: "Worksheet", *columns: Tuple[str, int]) -> None:
    """Write header columns and format their style and width."""
    for column, (value, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=column, value=value)
        cell.style = "header"
        ws.column_dimensions[cell.column_letter].width = width


def add_readme(wb: "Workbook") -> None:
    """Add a "README" sheet to `wb`."""
    ws = wb.create_sheet("README")

    _header(ws, ("Transport Data Commons (TDC) metadata", 72))
    ws["A3"] = README_TEXT


def add_attributes(wb: "Workbook", msd: "v21.MetadataStructureDefinition"):
    """Add an "Attributes" sheet to `wb` listing the metadata attributes from `msd`."""
    ws = wb.create_sheet("Attributes")

    _header(
        ws,
        ("Name", 20),  # "Element name" in Eurostat
        ("Description", 72),  # Not present in Eurostat
        ("ID", 20),  # "Element code" in Eurostat
    )

    for row, attribute in enumerate(msd.report_structure["ALL"].components, start=2):
        concept = attribute.concept_identity
        ws.cell(row=row, column=1, value=concept.name.localized_default()).style = "top"
        ws.cell(row=row, column=2, value=concept.description.localized_default())
        ws.cell(row=row, column=3, value=attribute.id).style = "top"


def add_template(wb: "Workbook", msd: "v21.MetadataStructureDefinition"):
    """Add a "TEMPLATE" sheet to `wb` with a metadata template."""
    ws = wb.create_sheet("TEMPLATE")

    _header(
        ws,
        ("Attribute name", 20),  # "Concept name" in Eurostat
        ("Value", 72),  # "Concept value" in Eurostat
    )

    for row, attribute in enumerate(msd.report_structure["ALL"].components, start=2):
        concept = attribute.concept_identity
        ws.cell(row=row, column=1, value=concept.name.localized_default()).style = "top"
        ws.cell(row=row, column=2, value="---")


def contains_data_for(mdr: "v21.MetadataReport", ref_area: str) -> bool:
    """Return :any:`True` if `mdr` contains data for `ref_area`.

    :any:`True` is returned if any of the following:

    1. The referenced data flow definition has an ID that starts with `ref_area`.
    2. The country's ISO 3166 alpha-2 code, alpha-3 code, official name, or common name
       appears in the value of the ``DATA_DESCR`` metadata attribute.


    Parameters
    ----------
    ref_area : str
        ISO 3166 alpha-2 code for a country. Passed to
        :meth:`pycountry.countries.lookup`.
    """
    country = countries.lookup(ref_area)

    if mdr.attaches_to.key_values["DATAFLOW"].obj.id.startswith(ref_area):  # type: ignore [union-attr]
        return True

    # Pattern to match in DATA_DESCR
    pat = re.compile(
        f"({country.alpha_2}|{country.alpha_3}|{country.name}|{country.common_name})"
    )
    for ra in mdr.metadata:
        assert hasattr(ra, "value")
        if ra.value_for.id == "DATA_DESCR" and pat.search(ra.value):
            return True

    return False


def generate_summary_html0(
    mds: "v21.MetadataSet", ref_area: str, path: "pathlib.Path"
) -> None:
    """Generate a summary report in HTML."""

    grouped = groupby(mds, key=partial(contains_data_for, ref_area=ref_area))

    env, common = get_jinja_env()

    path.write_text(
        env.get_template("template-metadata-0.html").render(
            ref_area=ref_area, matched=grouped[True], no_match=grouped[False], **common
        )
    )


def generate_summary_html1(
    mds: "v21.MetadataSet", ref_area: list[str], path: "pathlib.Path"
) -> None:
    data = {
        mdr.attaches_to.key_values["DATAFLOW"].obj.id: {  # type: ignore [union-attr]
            ra: contains_data_for(mdr, ra) for ra in ref_area
        }
        for mdr in mds.report
    }

    env, common = get_jinja_env()
    path.write_text(
        env.get_template("template-metadata-1.html").render(
            ref_area=ref_area, data=data, **common
        )
    )


@lru_cache
def get_cs_common() -> "common.ConceptScheme":
    """Create a shared concept scheme for the concepts referenced by dimensions.

    Concepts in this scheme have an annotation ``tdc-aka``, which is a list of alternate
    IDs recognized for the concept.
    """
    from . import get_agencyscheme

    as_ = get_agencyscheme()
    cs = common.ConceptScheme(id="CONCEPTS", maintainer=as_["TDCI"])

    cs.setdefault(
        id="CONFIDENTIALITY",
        annotations=[common.Annotation(id="tdc-aka", text=repr(["CONFIDIENTALITY"]))],
    )
    cs.setdefault(
        id="FUEL_TYPE",
        annotations=[common.Annotation(id="tdc-aka", text=repr(["Fuel type"]))],
    )
    cs.setdefault(
        id="REF_AREA",
        annotations=[
            common.Annotation(
                id="tdc-aka", text=repr(["Area", "Country", "Country code", "Region"])
            )
        ],
    )
    cs.setdefault(
        id="SERVICE",
        annotations=[common.Annotation(id="tdc-aka", text=repr(["FREIGHT_PASSENGER"]))],
    )
    cs.setdefault(
        id="TIME_PERIOD",
        annotations=[common.Annotation(id="tdc-aka", text=repr(["Time", "Year"]))],
    )

    return cs


@lru_cache
def get_jinja_env():
    """Return a Jinja2 environment for rendering templates."""
    from jinja2 import Environment, PackageLoader, select_autoescape

    # Create a Jinja environment
    env = Environment(
        loader=PackageLoader("transport_data", package_path="data/org"),
        extensions=["jinja2.ext.loopcontrols"],
        autoescape=select_autoescape(),
        trim_blocks=True,
        lstrip_blocks=True,
    )

    def _dfd_id(mdr):
        return mdr.attaches_to.key_values["DATAFLOW"].obj.id

    def _get_reported_attribute(mdr, id_):
        for ra in mdr.metadata:
            if ra.value_for.id == id_:
                return ra.value, ra.value_for
        return "—", None

    def _format_desc(dim):
        if desc := str(dim.get_annotation(id="tdc-description").text):
            return desc
        else:
            return "—"

    env.filters["dfd_id"] = _dfd_id
    env.filters["format_desc"] = _format_desc

    return env, dict(
        get_reported_attribute=_get_reported_attribute,
    )


def get_msd() -> "v21.MetadataStructureDefinition":
    from transport_data import STORE

    from . import get_agencyscheme

    TDCI = get_agencyscheme()["TDCI"]

    cs = common.ConceptScheme(id="METADATA_CONCEPTS", maintainer=TDCI)
    msd = v21.MetadataStructureDefinition(id="SIMPLE", version="1", maintainer=TDCI)
    rs = msd.report_structure["ALL"] = v21.ReportStructure(id="ALL")

    for id_, (name, description) in CONCEPTS.items():
        ci = cs.setdefault(id=id_, name=name, description=description)
        rs.getdefault(id_, concept_identity=ci)

    # NB Currently not supported by sdmx1; results in an empty XML collection
    STORE.write(msd)

    return msd


def getdefault(is_: "common.ItemScheme", other: "common.Item") -> "common.Item":
    """Return an item from `is_` matching `other`.

    Several methods are attempted to match `other` with an existing item:

    1. ID of `other` is identical to that of an existing item.
    2. Transformed ID of `other`—in upper case, " " replaced with "_" is identical to
       that of an existing item.
    3. ID of `other` is in the annotation ``tdc-aka``

    """
    # Exact match on ID or transformed ID
    for candidate in (other.id, other.id.upper().replace(" ", "_")):
        try:
            return is_[candidate]
        except KeyError:
            pass

    # Iterate over existing items
    for item in is_:
        # Eval the annotation "tdc-aka" for a list of alternate IDs for the item
        if aka := item.eval_annotation(id="tdc-aka"):
            if other.id in aka:
                return item

    # Still no match; create the item
    return is_.setdefault(id=other.id)


def groupby(
    mds: "v21.MetadataSet", key=Callable[["v21.MetadataReport"], Hashable]
) -> dict[Hashable, list["v21.MetadataReport"]]:
    """Group metadata reports in `mds` according to a `key` function.

    Similar to :func:`itertools.groupby`.
    """
    result: dict[Hashable, list["v21.MetadataReport"]] = defaultdict(list)
    for k, g in itertools.groupby(mds.report, key):
        result[k].extend(g)
    return result


def make_workbook(name="sample.xlsx") -> None:
    """Generate a :class:`openpyxl.Workbook` for exchange of metadata."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill

    wb = Workbook()

    # Delete the default sheet
    assert wb.active
    wb.remove(wb.active)

    # Create two named styles
    header = NamedStyle(name="header")
    header.fill = PatternFill("solid", fgColor="000000")
    header.font = Font(bold=True, color="ffffff", name="Calibri")
    wb.add_named_style(header)

    top = NamedStyle(name="top")
    top.alignment = Alignment(vertical="top", wrap_text=True)
    top.font = Font(name="Calibri")
    wb.add_named_style(top)

    # Generate the metadata structure definition
    msd = get_msd()

    # Add sheets
    add_readme(wb)
    add_attributes(wb, msd)
    add_template(wb, msd)

    # Save the file
    wb.save(name)


def parse_dimension(value: str) -> List[v21.Concept]:
    """Parse the description of a dimension from `value`.

    Supported values include:

    1. Multiple lines, with each line beginning "- ".
    2. A single line, with dimensions separated by ", ".
    3. A single dimension ID.
    """
    # Partial regular expressions for a dimension
    entry = r"(?P<id>.+?)(?: \((?P<description>[^\)]*)\))?"

    # Split `value` into potentially multiple values; separate dimension IDs from
    # description/annotation
    parts = []
    if matches := re.findall(f"^- {entry}$", value, flags=re.MULTILINE):
        # Multiple lines, with each line beginning "- "
        parts.extend(matches)
    elif matches := re.findall(f"{entry}(?:, |$)", value):
        # Single line, with dimensions separated by ", "
        # TODO Check behaviour if the ", " is within parentheses
        parts.extend(matches)
    elif 0 == len(parts):
        # None of the above → a single dimension label
        parts.append(value)

    # Convert to a list of Concept objects
    return [
        v21.Concept(id=id_, name=id_, description=description)
        for id_, description in parts
    ]


def read_workbook(
    path: "pathlib.Path",
) -> tuple["v21.MetadataSet", "v21.ConceptScheme"]:
    """Read a metadata set from the workbook at `path`."""
    from openpyxl import load_workbook

    wb = load_workbook(path)
    # Generate/retrieve the metadata structure definition
    msd = get_msd()

    mds = v21.MetadataSet(structured_by=msd)

    # Create a shared concept scheme for the concepts referenced by dimensions
    # TODO Collect, maybe with get_msd()
    cs_dims = get_cs_common()

    for ws in wb.worksheets:
        # Skip information sheets generated by methods in this file
        if ws.title in ("README", "Attributes", "TEMPLATE"):
            continue

        if r := read_worksheet(ws, msd, cs_dims):
            mds.report.append(r)

    return mds, cs_dims


def read_worksheet(
    ws: "Worksheet",
    msd: "v21.MetadataStructureDefinition",
    cs_dims: "v21.ConceptScheme",
) -> Optional["v21.MetadataReport"]:
    """Read a metadata report from the worksheet `ws`.

    Parameters
    ----------
    msd :
       Metadata structure definition.
    """
    # Mapping from names (not IDs) to MetadataAttributes
    mda_for_name = {
        str(mda.concept_identity.name): mda
        for mda in msd.report_structure["ALL"].components
    }

    # Create the target of the report: a data flow definition
    # TODO Expand this DFD and its associated data structure definition
    df_id_from_title = ws.title
    dfd = v21.DataflowDefinition(id=ws.title, maintainer=msd.maintainer)
    dsd = v21.DataStructureDefinition(id=ws.title, maintainer=msd.maintainer)
    dfd.structure = dsd

    # Create objects to associate the metadata report with the data flow definition
    iot = v21.IdentifiableObjectTarget()
    tok = v21.TargetObjectKey(
        key_values={"DATAFLOW": v21.TargetIdentifiableObject(value_for=iot, obj=dfd)}
    )

    # Create the report itself
    mdr = v21.MetadataReport()
    mdr.attaches_to = tok

    mda = None  # Reference to the MetaDataAttribute describing the current row
    dimension_concepts = []

    # Iterate over rows in the worksheet, skipping the first
    for row in ws.iter_rows(min_row=2):
        try:
            # Column B: value in the row
            ra_value = row[1].value

            if ra_value is None:
                continue
        except IndexError:
            log.warning(
                f"Sheet {df_id_from_title!r} has only < 2 columns in the first row; skip"
            )
            return None

        # Column A: name of the metadata attribute
        mda_name = row[0].value

        # Identify the MDA
        # NB if `mda_name` is none, then `mda` retains the value found on the previous
        #    row. This allows e.g. multiple rows to give values for DIMENSION
        # TODO Protect against other malformed data.
        mda = mda_for_name.get(str(mda_name), mda)

        if mda and mda.id == "DIMENSION":
            # Parse 1 or more dimension(s) and add to the DSD
            dimension_concepts.extend(parse_dimension(str(ra_value)))
        else:
            # Store as OtherNonEnumeratedAttributeValue
            # TODO Use EnumeratedAttributeValue, once code lists are available
            #      corresponding to dimensions
            ra = v21.OtherNonEnumeratedAttributeValue(
                value=str(ra_value), value_for=mda
            )

            # Attend the reported attribute to the report
            mdr.metadata.append(ra)

    # Basic checks
    df_id_from_cell = _get(mdr, "DATAFLOW")
    if not df_id_from_cell:
        log.warning(f"Sheet {df_id_from_title!r} does not identify a data flow; skip")
        return None

    update_dimension_descriptor(dsd, cs_dims, *dimension_concepts)

    return mdr


def _get(mdr: "v21.MetadataReport", mda_id: str) -> Optional[str]:
    """Retrieve from `mdr` the reported value of the metadata attribute `mda_id`."""
    for mda in mdr.metadata:
        if mda.value_for is not None and mda.value_for.id == mda_id:
            assert hasattr(mda, "value")  # Exclude ReportedAttribute without value attr
            return mda.value
    # No match
    return None


def summarize_metadataattribute(mds: "v21.MetadataSet", mda_id: str) -> None:
    """Summarize unique values appear in metadata for attribute `mda_id`."""
    value_id = defaultdict(set)

    for r in mds.report:
        value_id[_get(r, mda_id) or "MISSING"].add(_get(r, "DATAFLOW") or "MISSING")

    assert mds.structured_by
    mda = mds.structured_by.report_structure["ALL"].get(mda_id)

    print("\n\n" + uline(f"{mda}: {len(value_id)} unique values"))
    for value, df_ids in sorted(value_id.items()):
        print(f"{value}\n    " + " ".join(sorted(df_ids)))


def summarize_metadatareport(mdr: "v21.MetadataReport") -> None:
    lines = ["", uline("Metadata report")]

    # Retrieve references to the data flow and data structure
    dfd: v21.DataflowDefinition = mdr.attaches_to.key_values["DATAFLOW"].obj  # type: ignore [union-attr]
    dsd = dfd.structure

    # Summarize the data flow and data structure

    lines.extend(
        [f"Refers to {dfd!r}", f"  with structure {dsd!r}", "    with dimensions:"]
    )
    for dim in dsd.dimensions:
        line = f"    - {dim.id}:"
        if desc := str(dim.get_annotation(id="tdc-description").text):
            line += f" {desc!s}"
        else:
            line += " —"
        try:
            original_id = dim.get_annotation(id="tdc-original-id").text
            line += f" ('{original_id!s}' in input file)"
        except KeyError:
            pass
        lines.append(line)

    lines.append("")

    for ra in mdr.metadata:
        if ra.value_for.id == "DATAFLOW":
            continue
        assert hasattr(ra, "value")
        lines.append(f"{ra.value_for}: {ra.value}")

    print("\n".join(lines))


def summarize_metadataset(mds: "v21.MetadataSet") -> None:
    """Print a summary of the contents of `mds`."""
    print(f"Metadata set containing {len(mds.report)} metadata reports")

    summarize_metadataattribute(mds, "MEASURE")
    summarize_metadataattribute(mds, "DATA_PROVIDER")
    summarize_metadataattribute(mds, "UNIT_MEASURE")

    for r in mds.report:
        summarize_metadatareport(r)


def uline(text: str, char: str = "=") -> str:
    """Underline `text`."""
    return f"{text}\n{char * len(text)}"


def update_dimension_descriptor(
    dsd: "v21.DataStructureDefinition", cs_dims: "v21.ConceptScheme", *concepts
) -> None:
    """Update the DimensionDescriptor of `dsd` with `concepts`."""
    for dc in concepts:
        # Identify the concept in `cs_dims` with the same ID
        c = getdefault(cs_dims, dc)

        # Construct annotations
        anno = [common.Annotation(id="tdc-description", text=dc.description)]
        if c.id != dc.id:
            anno.append(common.Annotation(id="tdc-original-id", text=dc.id))

        dsd.dimensions.getdefault(id=c.id, concept_identity=c, annotations=anno)
