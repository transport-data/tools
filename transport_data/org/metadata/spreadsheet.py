"""Non-standard TDC Excel file format for collecting metadata."""

import logging
import re
from typing import TYPE_CHECKING, List, Optional, Tuple, cast

from sdmx.model import common, v21

if TYPE_CHECKING:
    import pathlib

    from openpyxl import Workbook
    from openpyxl.cell.cell import Cell
    from openpyxl.worksheet.worksheet import Worksheet

    from transport_data.util.sdmx import MAKeywords

log = logging.getLogger(__name__)


#: README text for the TDC metadata file format.
README_TEXT = """This file is an unofficial, prototype TDC format for metadata.
loosely imitates the Eurostat format. These files contain metadata (information
*about* data) based on the SDMX information model, but their layout (sheet
names, columns, etc.) is not specified by the SDMX standard, hence ‘unofficial’.

This file has the following sheets.

README
======

This sheet.

Attributes
==========

- One row per metadata attribute (or 'field').
- Columns for the name; description; and ID (short and machine-readable) of each
  attribute. See these descriptions to learn what to write for each attribute.

One or more additional sheets
=============================

- The name (or title) of each sheet corresponds to the identity (ID) of the data
  flow that is described by the metadata in that sheet.
- In Column A, the name of the metadata attribute. Each name MUST exactly
  match one appearing in the "Attributes" sheet. Some names MAY be omitted.
- In Column B, the actual metadata. These may be empty.

TEMPLATE
========

To add information about additional data flows not included in existing sheets
(above), you can copy and rename this sheet.
"""


def _header(ws: "Worksheet", *columns: Tuple[str, int]) -> None:
    """Write header columns and format their style and width."""
    for column, (value, width) in enumerate(columns, start=1):
        cell = cast("Cell", ws.cell(row=1, column=column, value=value))
        cell.style = "header"
        ws.column_dimensions[cell.column_letter].width = width


def add_readme(wb: "Workbook") -> None:
    """Add a "README" sheet to `wb`."""
    ws = wb.create_sheet("README")

    _header(ws, ("Transport Data Commons (TDC) metadata", 72))
    ws["A3"] = README_TEXT


def add_attributes(wb: "Workbook", msd: "v21.MetadataStructureDefinition"):
    """Add an "Attributes" sheet to `wb` listing the metadata attributes from `msd`."""
    ws = wb.create_sheet("Attributes")

    _header(
        ws,
        ("Name", 20),  # "Element name" in Eurostat
        ("Description", 72),  # Not present in Eurostat
        ("ID", 20),  # "Element code" in Eurostat
    )

    for row, attribute in enumerate(msd.report_structure["ALL"].components, start=2):
        concept = attribute.concept_identity
        ws.cell(row=row, column=1, value=concept.name.localized_default()).style = "top"
        ws.cell(row=row, column=2, value=concept.description.localized_default())
        ws.cell(row=row, column=3, value=attribute.id).style = "top"


def add_template(wb: "Workbook", msd: "v21.MetadataStructureDefinition"):
    """Add a "TEMPLATE" sheet to `wb` with a metadata template."""
    ws = wb.create_sheet("TEMPLATE")

    _header(
        ws,
        ("Attribute name", 20),  # "Concept name" in Eurostat
        ("Value", 72),  # "Concept value" in Eurostat
    )

    for row, attribute in enumerate(msd.report_structure["ALL"].components, start=2):
        concept = attribute.concept_identity
        ws.cell(row=row, column=1, value=concept.name.localized_default()).style = "top"
        ws.cell(row=row, column=2, value="---")


def getdefault(is_: "common.ItemScheme", other: "common.Item") -> "common.Item":
    """Return an item from `is_` matching `other`.

    Several methods are attempted to match `other` with an existing item:

    1. ID of `other` is identical to that of an existing item.
    2. Transformed ID of `other`—in upper case, " " replaced with "_" is identical to
       that of an existing item.
    3. ID of `other` is in the annotation ``tdc-aka``

    """
    # Exact match on ID or transformed ID
    for candidate in (other.id, other.id.upper().replace(" ", "_")):
        try:
            return is_[candidate]
        except KeyError:
            pass

    # Iterate over existing items
    for item in is_:
        # Eval the annotation "tdc-aka" for a list of alternate IDs for the item
        if aka := item.eval_annotation(id="tdc-aka"):
            if other.id in aka:
                return item

    # Still no match; create the item
    return is_.setdefault(id=other.id)


def make_workbook(name="sample.xlsx") -> None:
    """Generate a :class:`openpyxl.Workbook` for exchange of metadata."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill

    from transport_data.org.metadata import get_msd

    wb = Workbook()

    # Delete the default sheet
    assert wb.active
    wb.remove(wb.active)

    # Create two named styles
    header = NamedStyle(name="header")
    header.fill = PatternFill("solid", fgColor="000000")
    header.font = Font(bold=True, color="ffffff", name="Calibri")
    wb.add_named_style(header)

    top = NamedStyle(name="top")
    top.alignment = Alignment(vertical="top", wrap_text=True)
    top.font = Font(name="Calibri")
    wb.add_named_style(top)

    # Generate the metadata structure definition
    msd = get_msd()

    # Add sheets
    add_readme(wb)
    add_attributes(wb, msd)
    add_template(wb, msd)

    # Save the file
    wb.save(name)


def parse_dimension(value: str) -> List[v21.Concept]:
    """Parse the description of a dimension from `value`.

    Supported values include:

    1. Multiple lines, with each line beginning "- ".
    2. A single line, with dimensions separated by ", ".
    3. A single dimension ID.
    """
    # Partial regular expressions for a dimension
    entry = r"(?P<id>.+?)(?: \((?P<description>[^\)]*)\))?"

    # Split `value` into potentially multiple values; separate dimension IDs from
    # description/annotation
    parts = []
    if matches := re.findall(f"^- {entry}$", value, flags=re.MULTILINE):
        # Multiple lines, with each line beginning "- "
        parts.extend(matches)
    elif matches := re.findall(f"{entry}(?:, |$)", value):
        # Single line, with dimensions separated by ", "
        # TODO Check behaviour if the ", " is within parentheses
        parts.extend(matches)
    elif 0 == len(parts):  # pragma: no cover
        # None of the above → a single dimension label
        parts.append(value)

    # Convert to a list of Concept objects
    return [
        v21.Concept(id=id_.strip(), name=id_, description=description)
        for id_, description in parts
    ]


def read_workbook(
    path: "pathlib.Path",
) -> tuple["v21.MetadataSet", "v21.ConceptScheme"]:
    """Read a metadata set from the workbook at `path`."""
    from openpyxl import load_workbook

    from transport_data.org.metadata import get_cs_common, get_msd

    wb = load_workbook(path)
    # Generate/retrieve the metadata structure definition
    msd = get_msd()

    mds = v21.MetadataSet(structured_by=msd)

    # Create a shared concept scheme for the concepts referenced by dimensions
    # TODO Collect, maybe with get_msd()
    cs_dims = get_cs_common()

    for ws in wb.worksheets:
        # Skip information sheets generated by methods in this file
        if ws.title in ("README", "Attributes", "TEMPLATE"):
            continue

        if r := read_worksheet(ws, msd, cs_dims):
            mds.report.append(r)

    return mds, cs_dims


def read_worksheet(
    ws: "Worksheet",
    msd: "v21.MetadataStructureDefinition",
    cs_dims: "v21.ConceptScheme",
) -> Optional["v21.MetadataReport"]:
    """Read a metadata report from the worksheet `ws`.

    Parameters
    ----------
    msd :
       Metadata structure definition.
    """
    from transport_data import STORE
    from transport_data.org.metadata import _get

    # Mapping from names (not IDs) to MetadataAttributes
    mda_for_name = {
        str(mda.concept_identity.name): mda
        for mda in msd.report_structure["ALL"].components
    }

    # Create the target of the report: a data flow definition
    # TODO Expand this DFD and its associated data structure definition
    df_id_from_title = ws.title
    ma_args: "MAKeywords" = dict(
        id=ws.title, maintainer=msd.maintainer, version="1.0.0"
    )
    dfd = v21.DataflowDefinition(**ma_args)
    dsd = v21.DataStructureDefinition(**ma_args)
    dsd.measures.getdefault(id="OBS_VALUE")
    dsd.attributes.getdefault(id="COMMENT")
    dfd.structure = dsd

    # Create objects to associate the metadata report with the data flow definition
    iot = v21.IdentifiableObjectTarget()
    tok = v21.TargetObjectKey(
        key_values={"DATAFLOW": v21.TargetIdentifiableObject(value_for=iot, obj=dfd)}
    )

    # Create the report itself
    mdr = v21.MetadataReport()
    mdr.attaches_to = tok

    mda = None  # Reference to the MetaDataAttribute describing the current row
    dimension_concepts = []

    # Iterate over rows in the worksheet, skipping the first
    for row in ws.iter_rows(min_row=2):
        try:
            # Column B: value in the row
            ra_value = row[1].value

            if ra_value is None:
                continue
        except IndexError:  # pragma: no cover
            log.warning(
                f"Sheet {df_id_from_title!r} has only < 2 columns in the first row; skip"
            )
            return None

        # Column A: name of the metadata attribute
        mda_name = row[0].value

        # Identify the MDA
        # NB if `mda_name` is none, then `mda` retains the value found on the previous
        #    row. This allows e.g. multiple rows to give values for DIMENSION
        # TODO Protect against other malformed data.
        mda = mda_for_name.get(str(mda_name), mda)

        if mda and mda.id == "DIMENSION":
            # Parse 1 or more dimension(s) and add to the DSD
            dimension_concepts.extend(parse_dimension(str(ra_value)))
        else:
            # Store as OtherNonEnumeratedAttributeValue
            # TODO Use EnumeratedAttributeValue, once code lists are available
            #      corresponding to dimensions
            ra = v21.OtherNonEnumeratedAttributeValue(
                value=str(ra_value).strip(), value_for=mda
            )

            # Attend the reported attribute to the report
            mdr.metadata.append(ra)

    # Basic checks
    df_id_from_cell = _get(mdr, "DATAFLOW")
    if not df_id_from_cell:
        log.warning(f"Sheet {df_id_from_title!r} does not identify a data flow; skip")
        return None

    update_dimension_descriptor(dsd, cs_dims, *dimension_concepts)

    # Store the DSD and DFD
    STORE.set(dsd)
    STORE.set(dfd)

    return mdr


def update_dimension_descriptor(
    dsd: "v21.DataStructureDefinition", cs_dims: "v21.ConceptScheme", *concepts
) -> None:
    """Update the DimensionDescriptor of `dsd` with `concepts`."""
    for dc in concepts:
        # Identify the concept in `cs_dims` with the same ID
        c = getdefault(cs_dims, dc)

        # Construct annotations
        anno = [v21.Annotation(id="tdc-description", text=dc.description)]
        if c.id != dc.id:
            anno.append(v21.Annotation(id="tdc-original-id", text=dc.id))

        dsd.dimensions.getdefault(id=c.id, concept_identity=c, annotations=anno)
